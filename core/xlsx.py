from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from core.config import OUTPUT_DIR, TODAY

log = logging.getLogger(__name__)

_HDR_COLOR = "1F3864"
_ALT_COLOR = "EEF2F7"

_COLUMNS = [
    ("Date Found", 14), ("Municipality", 22), ("Region", 18),
    ("Contact Name", 22), ("Role / Title", 20), ("Email", 32),
    ("Phone", 18), ("Source URL", 30), ("Score", 10),
    ("Priority Tier", 18), ("Email Subject", 35), ("Email Draft", 55),
    ("Status", 14),
]


def save_to_xlsx(leads: list[dict]) -> Path | None:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor=_HDR_COLOR)
        data_font = Font(name="Arial", size=10)

        for col_idx, (label, width) in enumerate(_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[chr(64 + col_idx)].width = width
        ws.row_dimensions[1].height = 28

        for row_idx, lead in enumerate(leads, start=2):
            fill = PatternFill("solid", fgColor=(_ALT_COLOR if row_idx % 2 == 0 else "FFFFFF"))
            values = [
                TODAY,
                lead.get("municipality", ""),
                lead.get("region", ""),
                lead.get("contact_name", ""),
                lead.get("role", ""),
                lead.get("email", "") or "",
                lead.get("phone", "") or "",
                lead.get("source_url", ""),
                lead.get("_score", ""),
                lead.get("_tier", ""),
                lead.get("_subject", ""),
                lead.get("_draft", ""),
                "New",
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font      = data_font
                cell.fill      = fill
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 12))
            ws.row_dimensions[row_idx].height = 20

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:M{len(leads) + 1}"

        fname = OUTPUT_DIR / f"leads_{TODAY}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(fname)
        log.info(f"XLSX saved: {fname}")
        return fname
    except Exception as e:
        log.error(f"XLSX error: {e}")
        return None
