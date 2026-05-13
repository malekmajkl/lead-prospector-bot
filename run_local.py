#!/usr/bin/env python3
"""
run_local.py — CEO Assistant: Telegram Bot (lokální provoz)
============================================================
Hlavní vstupní bod. Spusťte tento soubor pro start bota.

Příkazy v Telegramu:
  /hledej [role] [oblast] [počet]  — spustí celý pipeline
  /sync                             — Gmail ↔ Sheets auto-sync (Sent + Replied)
  /status                           — statistiky z Lead Database
  /leady                            — dnešní nové leady
  /sent|/replied|/reviewed|/closed  — manuální update statusu leadu
  /help                             — nápověda
"""

import os, sys, json, time, logging, requests, base64, re
from datetime import date, datetime
from pathlib import Path
from email.mime.text import MIMEText

# ── Logging ───────────────────────────────────────────────────────────────────
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(message)s",
    level=logging.INFO,
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler("logs/bot.log", encoding="utf-8"),
    ],
)
log = logging.getLogger(__name__)

# ── Load .env ─────────────────────────────────────────────────────────────────
def load_env():
    env_path = Path(__file__).parent / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ.setdefault(k.strip(), v.strip())

load_env()

TOKEN     = os.getenv("TELEGRAM_TOKEN", "")
CHAT_ID   = os.getenv("TELEGRAM_CHAT_ID", "")
API_KEY   = os.getenv("ANTHROPIC_API_KEY", "")
GMAIL_TOK = os.getenv("GMAIL_TOKEN_JSON", "gmail_token.json")
SHEETS_ID = os.getenv("SHEETS_ID", "")
CREDS     = os.getenv("GOOGLE_CREDS_JSON", "./service_account.json")
BASE_URL  = f"https://api.telegram.org/bot{TOKEN}"
TODAY     = date.today().isoformat()
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

# ── Startup checks ────────────────────────────────────────────────────────────
errors = []
if not TOKEN:   errors.append("❌ TELEGRAM_TOKEN chybí v .env")
if not API_KEY: errors.append("❌ ANTHROPIC_API_KEY chybí v .env")
if errors:
    for e in errors: print(e)
    sys.exit(1)



# ══════════════════════════════════════════════════════════════════════════════
# SKILL ENGINE — loads SKILL.md files and drives Claude API calls
# ══════════════════════════════════════════════════════════════════════════════

SKILLS_DIR = Path(os.getenv("SKILLS_DIR", str(Path(__file__).parent / "skills")))

def load_skill(name: str) -> str:
    """Load a SKILL.md file — searches local skills/ folder and ~/.openclaw/workspace/skills/"""
    candidates = [
        SKILLS_DIR / name / "SKILL.md",
        Path(__file__).parent / "skills" / name / "SKILL.md",
        Path.home() / ".openclaw" / "workspace" / "skills" / name / "SKILL.md",
    ]
    for path in candidates:
        if path.exists():
            log.info(f"Loaded skill: {name} ({path})")
            return path.read_text(encoding="utf-8")
    log.warning(f"Skill '{name}' not found — using built-in fallback")
    return ""


def call_claude(system: str, user_message: str,
                tools: list = None, max_tokens: int = 2000) -> str:
    """Call Claude API. Retries 3x on rate limit with exponential backoff."""
    import anthropic
    client = anthropic.Anthropic(api_key=API_KEY)
    kwargs = dict(model="claude-sonnet-4-6", max_tokens=max_tokens,
                  system=system, messages=[{"role": "user", "content": user_message}])
    if tools:
        kwargs["tools"] = tools
    for attempt in range(3):
        try:
            response = client.messages.create(**kwargs)
            return "".join(b.text for b in response.content if hasattr(b, "text")).strip()
        except Exception as e:
            if "rate_limit" in str(e).lower() or "429" in str(e):
                wait = 65 * (attempt + 1)
                log.warning(f"Rate limit — waiting {wait}s (attempt {attempt+1}/3)")
                time.sleep(wait)
            else:
                raise
    log.error("Rate limit — all retries exhausted"); return ""


def parse_json_response(raw: str):
    """Robustly extract JSON from Claude response."""
    raw = re.sub(r"^```(?:json)?\s*", "", raw.strip())
    raw = re.sub(r"\s*```$", "", raw).strip()
    for pattern in [r"\[.*\]", r"\{.*\}"]:
        match = re.search(pattern, raw, re.DOTALL)
        if match:
            try: return json.loads(match.group(0))
            except json.JSONDecodeError: pass
    try: return json.loads(raw)
    except json.JSONDecodeError as e:
        log.error(f"JSON parse failed: {e}\nRaw: {raw[:200]}"); return None


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — LEAD RESEARCH (skill-driven)
# ══════════════════════════════════════════════════════════════════════════════

def research_leads(role: str, region: str, count: int,
                   exclude_municipalities: list = None) -> list:
    """Research leads using lead-researcher SKILL.md as system prompt."""
    skill = load_skill("lead-researcher")
    exclude_note = ""
    if exclude_municipalities:
        exclude_note = (f"\n\nIMPORTANT: Skip these municipalities (already in DB): "
                       f"{', '.join(exclude_municipalities)}. Find different ones.")
    system = (skill or "You are an expert at finding Czech municipal contacts.") + """

## CRITICAL OUTPUT REQUIREMENT
Return ONLY a valid JSON array — no markdown, no explanation, no backticks.
Each object must have: municipality, region, contact_name, role, email (or null),
phone (or null), source_url, language. Never invent contact details.
"""
    raw = call_claude(
        system=system,
        user_message=(f"Find {count} contacts with role '{role}' in '{region}', Czech Republic."
                      f"{exclude_note} Return ONLY JSON array with {count} entries."),
        tools=[{"type": "web_search_20250305", "name": "web_search"}],
        max_tokens=3000,
    )
    result = parse_json_response(raw)
    if isinstance(result, list):
        log.info(f"Research: {len(result)} leads found")
        return result[:count]
    return []


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — LEAD SCORING (skill-driven)
# ══════════════════════════════════════════════════════════════════════════════

def score_leads_batch(leads: list) -> list:
    """Score all leads in one Claude call using lead-scorer SKILL.md."""
    if not leads: return []
    skill = load_skill("lead-scorer")
    system = (skill or "Score leads 0-100: role seniority(30), data completeness(30), municipality size(25), source quality(15).") + """

## OUTPUT REQUIREMENT
Return ONLY a valid JSON array with ALL original fields plus:
  "_score": integer 0-100
  "_tier": "🔴 High Priority" | "🟡 Medium Priority" | "🟢 Low Priority" | "⚪ Deprioritise"
Return ALL leads in same order. No text, no backticks.
"""
    raw = call_claude(system=system, max_tokens=3000,
                      user_message=f"Score these {len(leads)} leads:\n{json.dumps(leads, ensure_ascii=False, indent=2)}")
    result = parse_json_response(raw)
    if isinstance(result, list) and len(result) == len(leads):
        log.info(f"Scored {len(result)} leads via skill")
        return sorted(result, key=lambda x: x.get("_score", 0), reverse=True)
    log.warning("Skill scoring failed — using built-in fallback")
    return sorted([_score_fallback(l) for l in leads], key=lambda x: x["_score"], reverse=True)


def _score_fallback(lead: dict) -> dict:
    ROLE_PTS = {"starosta":30,"starostka":30,"primátor":30,"místostarosta":25,
                "ředitel":25,"it ředitel":20,"vedoucí odboru":18,"tajemník":15,"referent":10}
    LARGE = {"kroměříž","zlín","vsetín","uherské hradiště","uherský brod","brno","ostrava","olomouc","plzeň","holešov"}
    MEDIUM = {"luhačovice","vizovice","valašské meziříčí","kunovice","staré město","uherský ostroh"}
    role_l = (lead.get("role") or "").lower()
    rpts = next((v for k,v in ROLE_PTS.items() if k in role_l), 5)
    email = lead.get("email") or ""
    dpts = (8 if any(x in email for x in ["info@","podatelna@","e-podatelna@"]) else 15) if "@" in email else 0
    if lead.get("phone"): dpts += 10
    if (lead.get("contact_name") or "").count(" ") >= 1: dpts += 5
    muni = (lead.get("municipality") or "").lower().strip()
    spts = 25 if muni in LARGE else 20 if muni in MEDIUM else 10
    src = (lead.get("source_url") or "").lower()
    srcpts = 15 if ".cz" in src or ".eu" in src else (7 if src else 2)
    total = rpts + dpts + spts + srcpts
    lead["_score"] = total
    lead["_tier"] = ("🔴 High Priority" if total >= 75 else "🟡 Medium Priority" if total >= 50
                     else "🟢 Low Priority" if total >= 25 else "⚪ Deprioritise")
    return lead


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — DRAFT EMAILS (skill-driven)
# ══════════════════════════════════════════════════════════════════════════════

def draft_emails_batch(leads: list) -> list:
    """Draft personalized emails for all leads using gmail-lead-drafter SKILL.md."""
    if not leads: return leads
    skill = load_skill("gmail-lead-drafter")
    system = (skill or "Draft personalized Czech cold-outreach emails for municipal contacts.") + """

## CEO PROFILE
Jméno: Jan Novák | Titul: Jednatel | Společnost: SolarObec s.r.o.
Email: jan.novak@solarobec.cz | Telefon: +420 777 123 456
Produkt: Energetické úspory, fotovoltaika a dotace z OPŽP pro obce.

## OUTPUT REQUIREMENT
Return ONLY a valid JSON array with ALL original fields plus:
  "_subject": personalized subject line
  "_draft": full email body (Czech, under 200 words, formal tone)
Return ALL leads. No text before or after. No backticks.
"""
    raw = call_claude(system=system, max_tokens=4000,
                      user_message=f"Draft emails for {len(leads)} leads:\n{json.dumps(leads, ensure_ascii=False, indent=2)}")
    result = parse_json_response(raw)
    if isinstance(result, list) and len(result) == len(leads):
        log.info(f"Drafted {len(result)} emails via skill")
        return result
    log.warning("Skill drafting failed — using built-in fallback")
    return [_draft_fallback(l) for l in leads]


def _draft_fallback(lead: dict) -> dict:
    name = lead.get("contact_name",""); muni = lead.get("municipality","")
    role_l = (lead.get("role") or "").lower()
    surname = name.split()[-1] if name else ""
    sal = (f"Vážená paní {surname}" if any(x in role_l for x in ["starostka","ředitelka","tajemnice"])
           else f"Vážený pane {surname}" if any(x in role_l for x in ["starosta","ředitel","tajemník"])
           else f"Vážená paní / Vážený pane {surname}")
    lead["_subject"] = f"Energetické úspory pro {muni} — SolarObec s.r.o."
    lead["_draft"] = (f"{sal},\n\ndovoluji si Vás oslovit jménem SolarObec s.r.o. "
                      f"Specializujeme se na energetické úspory a fotovoltaiku pro obce v ČR.\n\n"
                      f"Obcím jako {muni} pomáháme snižovat náklady na energie o 30–60 % ročně "
                      f"a zajišťujeme dotace z OPŽP a modernizačního fondu.\n\n"
                      f"Rád bych Vám vše krátce představil — stačí 20 minut.\n\n"
                      f"S pozdravem,\nJan Novák\nJednatel, SolarObec s.r.o.\n"
                      f"jan.novak@solarobec.cz | +420 777 123 456")
    return lead


# ══════════════════════════════════════════════════════════════════════════════
# TELEGRAM HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def tg_send(chat_id, text, parse_mode="Markdown"):
    for chunk in [text[i:i+4000] for i in range(0, len(text), 4000)]:
        requests.post(f"{BASE_URL}/sendMessage", json={
            "chat_id": chat_id, "text": chunk, "parse_mode": parse_mode,
        }, timeout=15)
        time.sleep(0.3)

def tg_typing(chat_id):
    requests.post(f"{BASE_URL}/sendChatAction",
                  json={"chat_id": chat_id, "action": "typing"}, timeout=5)

def tg_updates(offset=None):
    r = requests.get(f"{BASE_URL}/getUpdates",
                     params={"timeout": 30, "offset": offset}, timeout=40)
    return r.json().get("result", []) if r.ok else []


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — LEAD RESEARCH
# ══════════════════════════════════════════════════════════════════════════════

RESEARCH_SYSTEM = """Jsi expert na výzkum kontaktů pro municipální sektor v České republice.
Najdi požadované kontakty pomocí web_search na oficiálních stránkách obcí.

Výstup MUSÍ být POUZE validní JSON array — žádný text před ani za, žádné backticky.
Každý objekt musí mít přesně tato pole (null pokud nenajdeš):
{
  "municipality": "název obce",
  "region": "název kraje",
  "contact_name": "celé jméno",
  "role": "funkce/titul",
  "email": "email nebo null",
  "phone": "telefon nebo null",
  "source_url": "URL zdroje",
  "language": "CZ"
}
Nikdy nevymýšlej kontaktní údaje. Preferuj osobní email před obecným."""


def research_leads(role, region, count, exclude_municipalities=None):
    import anthropic
    client = anthropic.Anthropic(api_key=API_KEY)

    exclude_note = ""
    if exclude_municipalities:
        exclude_list = ", ".join(exclude_municipalities)
        exclude_note = f" Vynech tyto obce (již v databázi): {exclude_list}."

    # Retry with backoff on rate limit
    for attempt in range(3):
        try:
            response = client.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=2000,
                system=RESEARCH_SYSTEM,
                tools=[{"type": "web_search_20250305", "name": "web_search"}],
                messages=[{"role": "user", "content":
                    f"Najdi {count} kontaktů s rolí '{role}' v oblasti '{region}' v ČR."
                    f"{exclude_note} Vrať POUZE JSON array s {count} záznamy."}],
            )
            break
        except anthropic.RateLimitError:
            wait = 60 * (attempt + 1)
            log.warning(f"Rate limit — čekám {wait}s (pokus {attempt+1}/3)")
            time.sleep(wait)
    else:
        log.error("Rate limit — všechny pokusy vyčerpány")
        return []

    raw = "".join(b.text for b in response.content if hasattr(b, "text")).strip()
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw).strip()
    match = re.search(r"\[.*\]", raw, re.DOTALL)
    if match:
        try:
            leads = json.loads(match.group(0))
            if isinstance(leads, list):
                return leads[:count]
        except json.JSONDecodeError as e:
            log.error(f"JSON parse error: {e}")
    return []


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — SCORING
# ══════════════════════════════════════════════════════════════════════════════

ROLE_PTS = {
    "starosta": 30, "starostka": 30, "primátor": 30,
    "místostarosta": 25, "ředitel": 25, "ředitelka": 25,
    "it ředitel": 20, "it manager": 20, "vedoucí odboru": 18,
    "tajemník": 15, "tajemnice": 15, "referent": 10,
}
LARGE  = {"kroměříž","zlín","vsetín","uherské hradiště","otrokovice","holešov",
           "uherský brod","brno","ostrava","olomouc","plzeň"}
MEDIUM = {"luhačovice","vizovice","valašské meziříčí","rožnov pod radhoštěm",
          "uherský ostroh","kunovice","staré město","bystřice pod hostýnem"}

def score_lead(lead):
    role_l = (lead.get("role") or "").lower()
    rpts = next((v for k,v in ROLE_PTS.items() if k in role_l), 5)
    email = lead.get("email") or ""
    dpts = (8 if any(x in email for x in ["info@","podatelna@","e-podatelna@"]) else 15) if "@" in email else 0
    if lead.get("phone"): dpts += 10
    if (lead.get("contact_name") or "").strip().count(" ") >= 1: dpts += 5
    muni = (lead.get("municipality") or "").lower().strip()
    spts = 25 if muni in LARGE else 20 if muni in MEDIUM else 10
    src = (lead.get("source_url") or "").lower()
    srcpts = 15 if ".cz" in src or ".eu" in src else (7 if src else 2)
    total = rpts + dpts + spts + srcpts
    tier = ("🔴 High Priority" if total >= 75 else
            "🟡 Medium Priority" if total >= 50 else
            "🟢 Low Priority" if total >= 25 else "⚪ Deprioritise")
    return total, tier


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — DEDUPLICATION
# ══════════════════════════════════════════════════════════════════════════════

def deduplicate(leads):
    # Logic follows lead-deduplicator/SKILL.md (email > name+muni > phone priority)
    if not SHEETS_ID or not os.path.exists(CREDS):
        log.info("Sheets not configured — skipping dedup")
        return leads, []
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds  = Credentials.from_service_account_file(
            CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client = gspread.authorize(creds)
        ws     = client.open_by_key(SHEETS_ID).worksheet("Leads")
        rows   = ws.get_all_records()
        existing_emails = {r.get("Email","").lower().strip() for r in rows if r.get("Email")}
        existing_pairs  = {(r.get("Contact Name","").lower().strip(),
                            r.get("Municipality","").lower().strip()) for r in rows}
        new_leads, dupes = [], []
        for lead in leads:
            email = (lead.get("email") or "").lower().strip()
            name  = (lead.get("contact_name") or "").lower().strip()
            muni  = (lead.get("municipality") or "").lower().strip()
            if (email and email in existing_emails) or (name, muni) in existing_pairs:
                dupes.append(lead)
            else:
                new_leads.append(lead)
        log.info(f"Dedup: {len(new_leads)} new, {len(dupes)} duplicates")
        return new_leads, dupes
    except Exception as e:
        log.warning(f"Dedup error: {e}")
        return leads, []




def get_known_municipalities():
    """Load all municipality names already in Sheets DB — used to pre-exclude before search."""
    if not SHEETS_ID or not os.path.exists(CREDS):
        return set()
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds  = Credentials.from_service_account_file(
            CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client = gspread.authorize(creds)
        ws     = client.open_by_key(SHEETS_ID).worksheet(
                     os.getenv("SHEETS_TAB", "Leads"))
        rows   = ws.get_all_records()
        munis  = {r.get("Municipality","").strip() for r in rows if r.get("Municipality")}
        log.info(f"Pre-loaded {len(munis)} known municipalities from Sheets")
        return munis
    except Exception as e:
        log.warning(f"Could not pre-load municipalities: {e}")
        return set()

# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — DRAFT EMAIL
# ══════════════════════════════════════════════════════════════════════════════

def draft_email(lead):
    name   = lead.get("contact_name", "")
    muni   = lead.get("municipality", "")
    role_l = (lead.get("role") or "").lower()
    surname = name.split()[-1] if name else ""

    if any(x in role_l for x in ["starostka","ředitelka","tajemnice"]):
        salutation = f"Vážená paní {surname}"
    elif any(x in role_l for x in ["starosta","ředitel","tajemník","primátor"]):
        salutation = f"Vážený pane {surname}"
    else:
        salutation = f"Vážená paní / Vážený pane {surname}"

    subject = f"Energetické úspory pro {muni} — SolarObec s.r.o."
    body = (
        f"{salutation},\n\n"
        f"dovoluji si Vás oslovit jménem SolarObec s.r.o. "
        f"Specializujeme se na energetické úspory a fotovoltaiku pro obce a města v ČR.\n\n"
        f"Obcím podobné velikosti jako {muni} pomáháme snižovat náklady na energie "
        f"o 30–60 % ročně a zajišťujeme kompletní administraci dotací z OPŽP "
        f"a modernizačního fondu — bez zátěže na Vašem úřadu.\n\n"
        f"Rád bych Vám možnosti krátce představil — stačí 20 minut online nebo osobně.\n\n"
        f"S pozdravem,\n"
        f"Jan Novák\n"
        f"Jednatel, SolarObec s.r.o.\n"
        f"jan.novak@solarobec.cz | +420 777 123 456"
    )
    return subject, body


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — SAVE TO SHEETS
# ══════════════════════════════════════════════════════════════════════════════

COLUMNS = ["Date Found","Municipality","Region","Contact Name","Role / Title",
           "Email","Phone","Source URL","Language","Email Draft","Status",
           "CEO Notes","Score","Priority Tier"]

SHEET_HEADERS = ["Date Found","Municipality","Region","Contact Name","Role / Title",
                 "Email","Phone","Source URL","Language","Email Draft","Status",
                 "CEO Notes","Score","Priority Tier"]

def save_to_sheets(leads):
    # Logic follows sheets-lead-sync/SKILL.md (header check, append, status updates)
    if not SHEETS_ID or not os.path.exists(CREDS):
        return 0
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds  = Credentials.from_service_account_file(
            CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client = gspread.authorize(creds)
        ws     = client.open_by_key(SHEETS_ID).worksheet(
                     os.getenv("SHEETS_TAB", "Leads"))

        # Ensure header row exists — prevents column shift bug
        first_row = ws.row_values(1)
        if not first_row or first_row[0] != "Date Found":
            ws.insert_row(SHEET_HEADERS, index=1)
            log.info("Header row inserted into Sheets")

        saved = 0
        for lead in leads:
            row = [
                TODAY,
                lead.get("municipality",""),
                lead.get("region",""),
                lead.get("contact_name",""),
                lead.get("role",""),
                lead.get("email","") or "",
                lead.get("phone","") or "",
                lead.get("source_url",""),
                "CZ",
                lead.get("_draft",""),
                "New", "",
                str(lead.get("_score","")),
                lead.get("_tier",""),
            ]
            ws.append_row(row, value_input_option="USER_ENTERED",
                          table_range="A1")
            saved += 1
        log.info(f"Saved {saved} leads to Sheets")
        return saved
    except Exception as e:
        log.error(f"Sheets save error: {e}")
        return 0


# ══════════════════════════════════════════════════════════════════════════════
# STEP 6 — SAVE TO XLSX (local backup)
# ══════════════════════════════════════════════════════════════════════════════

def save_to_xlsx(leads):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "Leads"
        HDR = "1F3864"; ALT = "EEF2F7"
        cols = [("Date Found",14),("Municipality",22),("Region",18),
                ("Contact Name",22),("Role / Title",20),("Email",32),
                ("Phone",18),("Source URL",30),("Score",10),
                ("Priority Tier",18),("Email Subject",35),("Email Draft",55),("Status",14)]
        hf = Font(name="Arial",bold=True,color="FFFFFF",size=10)
        hfill = PatternFill("solid",fgColor=HDR)
        for i,(label,width) in enumerate(cols,1):
            c = ws.cell(row=1,column=i,value=label)
            c.font=hf; c.fill=hfill
            c.alignment=Alignment(horizontal="center",vertical="center")
            ws.column_dimensions[chr(64+i)].width=width
        ws.row_dimensions[1].height=28
        df = Font(name="Arial",size=10)
        for ri,lead in enumerate(leads,2):
            fill=PatternFill("solid",fgColor=(ALT if ri%2==0 else "FFFFFF"))
            for ci,val in enumerate([
                TODAY, lead.get("municipality",""), lead.get("region",""),
                lead.get("contact_name",""), lead.get("role",""),
                lead.get("email","") or "", lead.get("phone","") or "",
                lead.get("source_url",""), lead.get("_score",""),
                lead.get("_tier",""), lead.get("_subject",""),
                lead.get("_draft",""), "New"
            ],1):
                c=ws.cell(row=ri,column=ci,value=val)
                c.font=df; c.fill=fill
                c.alignment=Alignment(vertical="center",wrap_text=(ci==12))
            ws.row_dimensions[ri].height=20
        ws.freeze_panes="A2"
        ws.auto_filter.ref=f"A1:M{len(leads)+1}"
        fname=OUTPUT_DIR/f"leads_{TODAY}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(fname)
        log.info(f"XLSX saved: {fname}")
        return fname
    except Exception as e:
        log.error(f"XLSX error: {e}"); return None


# ══════════════════════════════════════════════════════════════════════════════
# STEP 7 — GMAIL DRAFTS
# ══════════════════════════════════════════════════════════════════════════════

def save_gmail_drafts(leads):
    tok = Path(GMAIL_TOK)
    if not tok.exists():
        log.warning(f"Gmail token not found: {tok}"); return 0
    try:
        from googleapiclient.discovery import build
        from google.oauth2.credentials import Credentials
        service = build("gmail","v1",credentials=Credentials.from_authorized_user_file(str(tok)))
    except Exception as e:
        log.error(f"Gmail auth: {e}"); return 0
    saved = 0
    for lead in leads:
        to_email = lead.get("email") or ""
        if not to_email or "@" not in to_email: continue
        try:
            msg = MIMEText(lead["_draft"],"plain","utf-8")
            msg["To"]=to_email; msg["Subject"]=lead["_subject"]
            raw=base64.urlsafe_b64encode(msg.as_bytes()).decode()
            service.users().drafts().create(userId="me",body={"message":{"raw":raw}}).execute()
            saved+=1; log.info(f"Draft saved → {to_email}")
        except Exception as e:
            log.error(f"Draft failed for {to_email}: {e}")
    return saved


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════

def run_pipeline(chat_id, role, region, count):
    try:
        sheets_active = SHEETS_ID and os.path.exists(CREDS)
        MAX_ROUNDS = 5          # max search rounds
        MAX_SEARCHED = 40       # stop after searching this many total leads

        tg_send(chat_id,
            f"🔍 *Spouštím pipeline...*\n"
            f"• Role: `{role}`\n• Oblast: `{region}`\n• Hledám: `{count}` nových\n\n"
            f"⏳ Prohledávám weby obcí, čekejte 1–2 minuty...")
        tg_typing(chat_id)

        # Pre-load known municipalities from Sheets BEFORE first search
        # This avoids duplicates from round 1 — no wasted API calls
        excluded_munis = get_known_municipalities()
        if excluded_munis:
            tg_send(chat_id, f"📋 Načteno *{len(excluded_munis)}* obcí z databáze — vyhnu se duplicitám od začátku.")

        all_new_leads  = []
        total_searched = 0
        round_num      = 0

        while len(all_new_leads) < count and round_num < MAX_ROUNDS and total_searched < MAX_SEARCHED:
            round_num += 1
            needed = count - len(all_new_leads)

            if round_num > 1:
                # Wait 65s between rounds — API limit is 30k tokens/min
                tg_send(chat_id,
                    f"🔄 *Kolo {round_num}* — hledám dalších `{needed}` nových leadů\n"
                    f"_(přeskočeno {len(excluded_munis)} již známých obcí)_\n"
                    f"⏳ Čekám 65s kvůli API limitu...")
                time.sleep(65)
                tg_typing(chat_id)

            # Research — exclude already known municipalities
            leads = research_leads(role, region, needed + 2,
                                   exclude_municipalities=list(excluded_munis))
            if not leads:
                tg_send(chat_id, "⚠️ Žádné další leady nenalezeny v této oblasti.")
                break

            total_searched += len(leads)

            # Score all leads via lead-scorer skill
            leads = score_leads_batch(leads)

            # Draft emails via gmail-lead-drafter skill
            leads = draft_emails_batch(leads)

            # Track all searched municipalities to exclude next round
            for lead in leads:
                excluded_munis.add(lead.get("municipality","").strip())

            # Dedup against DB
            new_leads, dupes = deduplicate(leads)

            if dupes and not new_leads:
                log.info(f"Kolo {round_num}: {len(dupes)} duplicit, zkouším dál...")
                continue

            all_new_leads.extend(new_leads)
            log.info(f"Kolo {round_num}: +{len(new_leads)} nových, celkem {len(all_new_leads)}/{count}")

        # Nothing found at all
        if not all_new_leads:
            tg_send(chat_id,
                f"📭 *Žádné nové leady*\n\n"
                f"Prohledáno {total_searched} obcí v `{region}` — "
                f"vše je již v databázi.\n\n"
                f"💡 Zkuste jinou oblast nebo roli.")
            return

        # Save results
        saved_sheets = save_to_sheets(all_new_leads)
        xlsx_path    = save_to_xlsx(all_new_leads)
        draft_count  = save_gmail_drafts(all_new_leads)

        # Compact Telegram summary
        n = len(all_new_leads)
        today_fmt = datetime.now().strftime("%-d. %-m. %Y · %H:%M")
        lines = [
            f"✅ *{n} nových leadů nalezeno*"
            + (f" _(po {round_num} kolech)_" if round_num > 1 else ""),
            f"🔍 `{role}` | `{region}` | {today_fmt}",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━",
        ]
        for lead in all_new_leads:
            tier_icon = lead["_tier"].split()[0]
            lines.append(
                f"{tier_icon} *{lead.get('municipality','?')}* — "
                f"{lead.get('contact_name','?')}"
            )
        lines += [
            f"━━━━━━━━━━━━━━━━━━━━━━━━━",
            f"💾 Sheets: *{saved_sheets}/{n}*  "
            f"📬 Gmail: *{draft_count}/{n}*  "
            f"📁 `{xlsx_path.name if xlsx_path else '—'}`",
            f"_Detaily v Excel/Sheets_",
        ]
        tg_send(chat_id, "\n".join(lines))

    except Exception as e:
        log.exception("Pipeline error")
        tg_send(chat_id, f"❌ Chyba:\n`{str(e)}`")


# ══════════════════════════════════════════════════════════════════════════════
# COMMAND DISPATCHER
# ══════════════════════════════════════════════════════════════════════════════

WELCOME = """
👋 *CEO Assistant je připravený!*

*Vyhledávání leadů:*
`/hledej [role] [oblast] [počet]`
→ `/hledej starostové Kroměříž 5`

*Aktualizace statusu:*
`/sent [obec nebo příjmení]`
`/replied [obec nebo příjmení]`
`/reviewed [obec nebo příjmení]`
`/closed [obec nebo příjmení]`
→ `/sent holešov`  nebo  `/replied kubáník`

*Synchronizace:*
`/sync` — Gmail ↔ Sheets auto-sync (Sent + Replied)

*Přehled:*
`/status` — statistiky z databáze
`/leady`  — dnešní nové leady
`/help`   — nápověda
"""

def parse_hledej(text):
    parts = text.strip().split()
    count = 5
    if parts and parts[-1].isdigit():
        count = min(int(parts.pop()), 10)
    region_kw = ["kraj","kroměříž","zlín","brno","ostrava","olomouc",
                 "vsetín","jihlava","hradec","pardubice","hradiště","brod"]
    split_at = next((i for i,p in enumerate(parts) if any(k in p.lower() for k in region_kw)), None)
    if split_at and split_at > 0:
        role = " ".join(parts[:split_at])
        region = " ".join(parts[split_at:])
    elif len(parts) >= 2:
        role = parts[0]
        region = " ".join(parts[1:])
    else:
        role = " ".join(parts) or "starosta"
        region = "Česká republika"
    return role.strip(), region.strip(), count

def handle_status(chat_id):
    if not SHEETS_ID or not os.path.exists(CREDS):
        tg_send(chat_id,"⚠️ Google Sheets není připojeno.\nNastavte `SHEETS_ID` a `GOOGLE_CREDS_JSON` v `.env`.")
        return
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds  = Credentials.from_service_account_file(
            CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client = gspread.authorize(creds)
        rows   = client.open_by_key(SHEETS_ID).worksheet("Leads").get_all_records()
        counts = {}
        for r in rows:
            s = r.get("Status","").strip()
            counts[s] = counts.get(s,0)+1
        high   = sum(1 for r in rows if "High" in r.get("Priority Tier",""))
        medium = sum(1 for r in rows if "Medium" in r.get("Priority Tier",""))
        tg_send(chat_id,
            f"📊 *Lead Database — Statistiky*\n\n"
            f"• Celkem: *{len(rows)}*\n"
            f"• 🆕 New: *{counts.get('New',0)}*\n"
            f"• 👁 Reviewed: *{counts.get('Reviewed',0)}*\n"
            f"• 📤 Sent: *{counts.get('Sent',0)}*\n"
            f"• 💬 Replied: *{counts.get('Replied',0)}*\n"
            f"• ✅ Closed: *{counts.get('Closed',0)}*\n\n"
            f"• 🔴 High Priority: *{high}*\n"
            f"• 🟡 Medium Priority: *{medium}*")
    except Exception as e:
        tg_send(chat_id, f"❌ Chyba při načítání Sheets:\n`{e}`")


def find_lead(query: str) -> list[dict]:
    """
    Fuzzy search in Sheets — find leads matching query string.
    Matches against Municipality and Contact Name (case-insensitive, partial).
    Returns list of matching row dicts with added _row_index key.
    """
    if not SHEETS_ID or not os.path.exists(CREDS):
        return []
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds  = Credentials.from_service_account_file(
            CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client = gspread.authorize(creds)
        ws     = client.open_by_key(SHEETS_ID).worksheet(
                     os.getenv("SHEETS_TAB", "Leads"))
        rows   = ws.get_all_records()
        q      = query.lower().strip()

        # Remove czech diacritics for fuzzy matching
        import unicodedata
        def strip_diacritics(s):
            return ''.join(c for c in unicodedata.normalize('NFD', s)
                           if unicodedata.category(c) != 'Mn').lower()

        q_norm = strip_diacritics(q)
        matches = []
        for i, row in enumerate(rows, start=2):  # row 1 = header
            muni = strip_diacritics(row.get("Municipality", ""))
            name = strip_diacritics(row.get("Contact Name", ""))
            if q_norm in muni or q_norm in name:
                row["_row_index"] = i
                matches.append(row)
        return matches
    except Exception as e:
        log.error(f"find_lead error: {e}")
        return []


STATUS_LABELS = {
    "sent":     "Sent",
    "replied":  "Replied",
    "reviewed": "Reviewed",
    "closed":   "Closed",
    "new":      "New",
}

STATUS_ICONS = {
    "New":      "🆕",
    "Reviewed": "👁",
    "Sent":     "📤",
    "Replied":  "💬",
    "Closed":   "✅",
}

def update_lead_status_in_sheets(row_index: int, status: str, notes: str = "") -> bool:
    """Update Status (col K = 11) and optionally Notes (col L = 12) by row index."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds  = Credentials.from_service_account_file(
            CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client = gspread.authorize(creds)
        ws     = client.open_by_key(SHEETS_ID).worksheet(
                     os.getenv("SHEETS_TAB", "Leads"))
        ws.update_cell(row_index, 11, status)  # col K = Status
        if notes:
            ws.update_cell(row_index, 12, notes)  # col L = CEO Notes
        log.info(f"Row {row_index} status → {status}")
        return True
    except Exception as e:
        log.error(f"update_lead_status error: {e}")
        return False


def handle_update_status(chat_id: str, cmd: str, args: str):
    """
    Handle /sent, /replied, /reviewed, /closed, /new commands.
    cmd  = e.g. "sent"
    args = search query e.g. "holešov" or "kubáník"
    """
    if not args:
        tg_send(chat_id,
            f"⚠️ Zadej obec nebo příjmení.\n\nPříklad: `/{cmd} holešov`  nebo  `/{cmd} kubáník`")
        return

    if not SHEETS_ID or not os.path.exists(CREDS):
        tg_send(chat_id, "⚠️ Google Sheets není připojeno.")
        return

    new_status = STATUS_LABELS[cmd]
    icon       = STATUS_ICONS[new_status]
    matches    = find_lead(args)

    if not matches:
        tg_send(chat_id,
            f"❌ Lead *'{args}'* nenalezen v databázi.\n\nZkus jiný výraz nebo `/leady` pro výpis dnešních leadů.")
        return

    if len(matches) == 1:
        # Exact single match — update immediately
        lead = matches[0]
        ok   = update_lead_status_in_sheets(lead["_row_index"], new_status)
        if ok:
            tg_send(chat_id,
                f"{icon} *Status aktualizován*\n\n🏛️ {lead['Municipality']} — {lead['Contact Name']}\n📧 {lead.get('Email','—')}\nStatus: *{new_status}*")
        else:
            tg_send(chat_id, "❌ Nepodařilo se aktualizovat Sheets. Zkus znovu.")

    else:
        # Multiple matches — ask user to be more specific
        lines = [
            f"🔍 Nalezeno *{len(matches)}* výsledků pro `{args}`:",
            f"Upřesni dotaz:\n",
        ]
        for m in matches[:8]:  # max 8 suggestions
            old_status = m.get("Status", "?")
            lines.append(
                f"• `/{cmd} {m['Municipality'].lower()}` "
                f"— {m['Contact Name']} ({old_status})"
            )
        if len(matches) > 8:
            lines.append(f"_...a {len(matches)-8} dalších_")
        tg_send(chat_id, "\n".join(lines))


def handle_leady(chat_id: str):
    """Show today's new leads from Sheets."""
    if not SHEETS_ID or not os.path.exists(CREDS):
        tg_send(chat_id, "⚠️ Google Sheets není připojeno.")
        return
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        creds  = Credentials.from_service_account_file(
            CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        client = gspread.authorize(creds)
        rows   = client.open_by_key(SHEETS_ID).worksheet(
                     os.getenv("SHEETS_TAB","Leads")).get_all_records()
        today_leads = [r for r in rows
                       if str(r.get("Date Found","")) == TODAY
                       and r.get("Status","") == "New"]
        if not today_leads:
            tg_send(chat_id, f"📭 Dnes ({TODAY}) žádné nové leady.")
            return
        lines = [f"📋 *Dnešní nové leady ({len(today_leads)})*\n"]
        for l in today_leads:
            lines.append(
                f"🏛️ *{l.get('Municipality','?')}* — {l.get('Contact Name','?')}\n   {l.get('Priority Tier','')} · Score: {l.get('Score','—')}\n   📧 `{l.get('Email','—')}`\n"
            )
        tg_send(chat_id, "\n".join(lines))
    except Exception as e:
        tg_send(chat_id, f"❌ Chyba:\n`{e}`")




def handle_sync(chat_id: str):
    """
    Auto-sync Gmail ↔ Sheets:
    1. Check Gmail Sent → leads whose draft was sent → Status = Sent
    2. Check Gmail Inbox → replies from lead emails → Status = Replied
    3. Update Sheets for all matches
    """
    if not SHEETS_ID or not os.path.exists(CREDS):
        tg_send(chat_id, "⚠️ Google Sheets není připojeno.")
        return
    if not os.path.exists(GMAIL_TOK):
        tg_send(chat_id, "⚠️ Gmail token nenalezen.")
        return

    tg_send(chat_id, "🔄 *Synchronizuji Gmail ↔ Sheets...*\n⏳ Čekejte prosím...")

    try:
        # ── Gmail setup ───────────────────────────────────────────────────────
        from googleapiclient.discovery import build
        from google.oauth2.credentials import Credentials as GCreds
        gmail = build("gmail", "v1",
                      credentials=GCreds.from_authorized_user_file(GMAIL_TOK))

        # ── Sheets setup ──────────────────────────────────────────────────────
        import gspread
        from google.oauth2.service_account import Credentials as SAcreds
        sa    = SAcreds.from_service_account_file(
                    CREDS, scopes=["https://www.googleapis.com/auth/spreadsheets"])
        gc    = gspread.authorize(sa)
        ws    = gc.open_by_key(SHEETS_ID).worksheet(os.getenv("SHEETS_TAB","Leads"))
        rows  = ws.get_all_records()

        # Build email → row_index map from Sheets (only New/Reviewed/Sent leads)
        email_to_row = {}
        for i, row in enumerate(rows, start=2):
            email = row.get("Email","").strip().lower()
            status = row.get("Status","").strip()
            if email and status in ("New","Reviewed","Sent"):
                email_to_row[email] = {"row": i, "status": status,
                                        "name": row.get("Contact Name",""),
                                        "muni": row.get("Municipality","")}

        if not email_to_row:
            tg_send(chat_id, "ℹ️ Žádné aktivní leady v Sheets k synchronizaci.")
            return

        sent_updates    = []
        replied_updates = []

        # ── 1. Check Gmail SENT ───────────────────────────────────────────────
        # Get messages from Sent folder, last 30 days
        sent_result = gmail.users().messages().list(
            userId="me", labelIds=["SENT"], maxResults=100,
            q="newer_than:30d"
        ).execute()
        sent_messages = sent_result.get("messages", [])

        for msg_ref in sent_messages:
            msg = gmail.users().messages().get(
                userId="me", id=msg_ref["id"], format="metadata",
                metadataHeaders=["To","Subject"]
            ).execute()
            headers = {h["name"].lower(): h["value"]
                       for h in msg.get("payload",{}).get("headers",[])}
            to_raw = headers.get("to","")
            # Extract email address from "Name <email>" format
            import re as _re
            match = _re.search(r"<(.+?)>", to_raw)
            to_email = (match.group(1) if match else to_raw).strip().lower()

            if to_email in email_to_row:
                lead = email_to_row[to_email]
                if lead["status"] in ("New","Reviewed"):
                    sent_updates.append({
                        "row": lead["row"], "email": to_email,
                        "name": lead["name"], "muni": lead["muni"],
                        "subject": headers.get("subject","")
                    })

        # ── 2. Check Gmail INBOX for replies ─────────────────────────────────
        # Look for emails FROM any lead address
        lead_emails = list(email_to_row.keys())
        # Gmail search: from:(email1 OR email2 OR ...)
        # Batch in groups of 10 to avoid query length limit
        chunk_size = 10
        inbox_messages = []
        for i in range(0, min(len(lead_emails), 50), chunk_size):
            chunk = lead_emails[i:i+chunk_size]
            query = "from:(" + " OR ".join(chunk) + ") newer_than:60d"
            result = gmail.users().messages().list(
                userId="me", labelIds=["INBOX"], maxResults=50, q=query
            ).execute()
            inbox_messages.extend(result.get("messages", []))

        for msg_ref in inbox_messages:
            msg = gmail.users().messages().get(
                userId="me", id=msg_ref["id"], format="metadata",
                metadataHeaders=["From","Subject"]
            ).execute()
            headers = {h["name"].lower(): h["value"]
                       for h in msg.get("payload",{}).get("headers",[])}
            from_raw = headers.get("from","")
            import re as _re
            match = _re.search(r"<(.+?)>", from_raw)
            from_email = (match.group(1) if match else from_raw).strip().lower()

            if from_email in email_to_row:
                lead = email_to_row[from_email]
                replied_updates.append({
                    "row": lead["row"], "email": from_email,
                    "name": lead["name"], "muni": lead["muni"],
                    "subject": headers.get("subject","")
                })

        # ── 3. Apply updates to Sheets ────────────────────────────────────────
        # Replied takes priority over Sent
        replied_rows = {u["row"] for u in replied_updates}

        updated_sent    = 0
        updated_replied = 0

        for u in sent_updates:
            if u["row"] not in replied_rows:  # don't downgrade Replied → Sent
                ws.update_cell(u["row"], 11, "Sent")
                updated_sent += 1
                log.info(f"Sync: {u['muni']} → Sent")

        for u in replied_updates:
            ws.update_cell(u["row"], 11, "Replied")
            updated_replied += 1
            log.info(f"Sync: {u['muni']} → Replied")

        # ── 4. Build Telegram report ──────────────────────────────────────────
        total = updated_sent + updated_replied
        if total == 0:
            tg_send(chat_id,
                "✅ *Sync dokončen — žádné změny*\n\n"
                f"Zkontrolováno {len(sent_messages)} odeslaných emailů "
                f"a {len(inbox_messages)} odpovědí.\n"
                "_Vše je aktuální._")
            return

        lines = [
            f"✅ *Sync dokončen — {total} aktualizací*",
            f"━━━━━━━━━━━━━━━━━━━━━━━━━",
        ]
        if updated_sent:
            lines.append(f"\n📤 *Odesláno ({updated_sent}):*")
            for u in sent_updates:
                if u["row"] not in replied_rows:
                    lines.append(f"  📤 {u['muni']} — {u['name']}")
        if updated_replied:
            lines.append(f"\n💬 *Odpověděli ({updated_replied}):*")
            for u in replied_updates:
                lines.append(f"  💬 {u['muni']} — {u['name']}")
        lines += [
            f"\n━━━━━━━━━━━━━━━━━━━━━━━━━",
            f"_Sheets aktualizován · {datetime.now().strftime('%-d.%-m. %H:%M')}_"
        ]
        tg_send(chat_id, "\n".join(lines))

    except Exception as e:
        log.exception("Sync error")
        tg_send(chat_id, f"❌ Chyba při synchronizaci:\n`{str(e)}`")

def dispatch(message):
    chat_id = str(message["chat"]["id"])
    text    = message.get("text","").strip()
    if CHAT_ID and chat_id != str(CHAT_ID):
        tg_send(chat_id,"⛔ Přístup odepřen."); return
    log.info(f"[{chat_id}] {text[:80]}")
    if text.startswith("/start") or text.startswith("/help"):
        tg_send(chat_id, WELCOME)
    elif text.startswith("/hledej"):
        args = text[len("/hledej"):].strip()
        if not args:
            tg_send(chat_id,"⚠️ Zadej roli a oblast.\n\nPříklad: `/hledej starostové Kroměříž 5`")
            return
        role, region, count = parse_hledej(args)
        run_pipeline(chat_id, role, region, count)
    elif text.startswith("/status"):
        handle_status(chat_id)
    elif text.startswith("/leady"):
        handle_leady(chat_id)
    elif text.startswith("/sync"):
        handle_sync(chat_id)
    elif any(text.startswith(f"/{cmd}") for cmd in STATUS_LABELS):
        # /sent, /replied, /reviewed, /closed, /new
        parts = text.split(None, 1)
        cmd   = parts[0].lstrip("/").lower()
        args  = parts[1].strip() if len(parts) > 1 else ""
        handle_update_status(chat_id, cmd, args)
    else:
        tg_send(chat_id,
            "❓ Neznámý příkaz. Zkus `/help`.\n\n"
            "Rychlé tipy:\n"
            "`/hledej starostové Kroměříž 5`\n"
            "`/sent holešov`\n"
            "`/replied kubáník`")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN LOOP
# ══════════════════════════════════════════════════════════════════════════════

def main():
    sheets_ok = SHEETS_ID and os.path.exists(CREDS)
    print("\n" + "="*50)
    print("  CEO Assistant Bot")
    print(f"  {datetime.now().strftime('%d.%m.%Y %H:%M')}")
    print("="*50)
    print(f"✅ Telegram: připojeno")
    print(f"✅ Anthropic API: připojeno")
    print(f"{'✅' if sheets_ok else '⚠️ '} Google Sheets: {'připojeno' if sheets_ok else 'nepřipojeno'}")
    print(f"{'✅' if Path(GMAIL_TOK).exists() else '⚠️ '} Gmail: {'připojeno' if Path(GMAIL_TOK).exists() else 'nepřipojeno'}")
    print("\n💬 Napište /hledej v Telegramu")
    print("   Zastavení: Ctrl+C\n")

    offset = None
    while True:
        try:
            for u in tg_updates(offset):
                offset = u["update_id"] + 1
                if "message" in u:
                    dispatch(u["message"])
        except KeyboardInterrupt:
            print("\n👋 Bot zastaven."); break
        except requests.exceptions.ConnectionError:
            log.warning("Connection error — retry in 5s"); time.sleep(5)
        except Exception as e:
            log.exception(f"Polling error: {e}"); time.sleep(3)

if __name__ == "__main__":
    main()